#!/usr/bin/env python3
"""Safe fictional daily-control updater. Review before use."""
import argparse, csv, os, shutil
from datetime import datetime
from openpyxl import load_workbook

p=argparse.ArgumentParser(); p.add_argument("--input",required=True); p.add_argument("--workbook",required=True); p.add_argument("--output",required=True); p.add_argument("--dry-run",action="store_true"); args=p.parse_args()
wb=load_workbook(args.workbook); required={"Management_Control","Review_Log"}; missing=required-set(wb.sheetnames)
if missing: raise SystemExit(f"Missing sheets: {sorted(missing)}")
rows=list(csv.DictReader(open(args.input,encoding="utf-8")))
ws=wb["Management_Control"]; keys={(ws.cell(r,1).value,ws.cell(r,2).value):r for r in range(2,ws.max_row+1)}
changes=[]
for item in rows:
    key=(item["Month"],item["Measure"]); row=keys.get(key)
    if not row: continue
    old=ws.cell(row,4).value; new=float(item["Actual"])
    if old!=new: changes.append((row,old,new))
if args.dry_run:
    print(f"DRY RUN: {len(changes)} proposed changes"); [print(x) for x in changes]; raise SystemExit(0)
backup=args.workbook+"."+datetime.now().strftime("%Y%m%d-%H%M%S")+".bak"; shutil.copy2(args.workbook,backup)
for row,old,new in changes: ws.cell(row,4).value=new
log=wb["Review_Log"]; log.append([datetime.now(),"Daily update",f"{len(changes)} actual values updated",os.getenv("USER","training-user"),"","Pending review"])
os.makedirs(os.path.dirname(args.output) or ".",exist_ok=True); wb.save(args.output); print(args.output)
